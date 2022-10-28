#Dans ce script j'initialise une classe pour créer des cartes et leur règles, c'est vraiment inutile, je pourrais tout gérer avec juste un dictionnaire et une liste, je cherche à manipuler au mieux les classes

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import random
import csv

class Cartestats(object):
    """Classe permettant de créer des cartes personnalisées"""
    def __init__(self, nom, rarete, attaque, defense, agilite, intelligence, vitesse):
        self.nom = nom
        self.rarete = rarete
        self.attaque = attaque
        self.defense = defense
        self.agilite = agilite
        self.intelligence = intelligence
        self.vitesse = vitesse

    def __str__(self):
        return ("Nom: {}, Rareté: {}, Attaque: {}, Defense: {}, Agilité: {}, Intelligence: {}, Vitesse: {}".format(self.nom, self.rarete, self.attaque, self.defense, self.agilite, self.intelligence, self.vitesse))

    def __repr__(self):
        return ("Nom: {}, Rareté: {}, Attaque: {}, Defense: {}, Agilité: {}, Intelligence: {}, Vitesse: {}".format(self.nom, self.rarete, self.attaque, self.defense, self.agilite, self.intelligence, self.vitesse))

    def __gt__(self, other, statistique):
        """Définition des règles de grandeur"""
        if statistique == "attaque":
            if self.attaque > other.attaque:
                return True
            else:
                return False

        if statistique == "defense":
            if self.defense > other.defense:
                return True
            else:
                return False

        if statistique == "agilite":
            if self.agilite > other.agilite:
                return True
            else:
                return False       

        if statistique == "intelligence":
            if self.intelligence > other.intelligence:
                return True
            else:
                return False 

        if statistique == "vitesse":
            if self.vitesse > other.vitesse:
                return True
            else:
                return False

    def __ge__(self, other, statistique):
        """Permet de régir les égalités"""

        if statistique == "attaque":
            if self.attaque == other.attaque:
                return True
            else:
                return False

        if statistique == "defense":
            if self.defense == other.defense:
                return True
            else:
                return False

        if statistique == "agilite":
            if self.agilite == other.agilite:
                return True
            else:
                return False       

        if statistique == "intelligence":
            if self.intelligence == other.intelligence:
                return True
            else:
                return False 

        if statistique == "vitesse":
            if self.vitesse == other.vitesse:
                return True
            else:
                return False

class Jeudecartestats(object):
    """Classe permettant de jouer à différents modes de jeux à l'aide d'un paquet de carte"""
    def __init__(self):
        paquet = importtsv()
        self.paquet = paquet

    def melange(self, joueur):
        """Fonction permettant de mélanger plusieurs decks"""
        for n in joueur:
            random.shuffle(joueur[n])
            
        return joueur


    def bataillesimple(self):
        """Fonction permettant d'initialiser une bataille simple à partir de 2 joueurs"""
        joueur = {}
        listejoueurs = []
        repeat = True
        while repeat:
            try :
                newjoueur = str(input("Entrez le nom d'un nouveau joueur, laisser vide pour arrêter\n"))
            except:
                print("Ce n'est pas valide")
                
            if newjoueur != "":
                listejoueurs.append(newjoueur)
            elif len(listejoueurs) > 2:
                repeat = False
            else:
                print("Il faut au moins 2 joueurs\n")

        for n in listejoueurs:
            random.shuffle(self.paquet)
            joueur[n] = self.paquet
        print(joueur)
        return joueur




def importtsv():
    """Fonction permettant d'extraire les données des jeux vers des fichiers. Retourne un dictionnaire avec la liste de carte pour chaque rareté."""
    #On declare les différents types de cartes pouvant être appelées
    unique = []
    rare = []
    peucommun = []
    commun = []
    paquetdecarte = {}
    listepaquetdecarte = []
    with open("listedescartes.tsv", newline='', encoding="utf-8") as f:
        reader = csv.reader(f, delimiter = '\t', lineterminator="\n")
        for row in reader:
            if row[1] == "Commun" :
                newcarte = Cartestats(row[0], row[1], row[2], row[3], row[4], row[5], row[6])
                commun.append(newcarte)
                listepaquetdecarte.append(newcarte)
            elif row[1] == "Peu Commun":
                newcarte = Cartestats(row[0], row[1], row[2], row[3], row[4], row[5], row[6])
                peucommun.append(newcarte)
                listepaquetdecarte.append(newcarte)
            elif row[1] == "Rare":
                newcarte = Cartestats(row[0], row[1], row[2], row[3], row[4], row[5], row[6])
                rare.append(newcarte)
                listepaquetdecarte.append(newcarte)
            elif row[1] == "Unique":
                newcarte = Cartestats(row[0], row[1], row[2], row[3], row[4], row[5], row[6])
                unique.append(newcarte)
                listepaquetdecarte.append(newcarte)
            else:
                continue
    
    #On ajoute les listes de cartes dans leur dico respectifs et on retourne 
    paquetdecarte["Commun"], paquetdecarte["Peu Commun"], paquetdecarte["Rare"], paquetdecarte["Unique"] = commun, peucommun, rare, unique
    print(paquetdecarte["Commun"][0])
    print(paquetdecarte)
    return listepaquetdecarte



def importxlsx():
    """Fonction permettant d'extraire les données des jeux vers des fichiers. Retourne un dictionnaire avec la liste de carte pour chaque rareté."""
    #On declare les différents types de cartes pouvant être appelées
    unique = []
    rare = []
    peucommun = []
    commun = []
    paquetdecarte = {}
    listepaquetdecarte = []

    #On ouvre le fichier excel (google sheets) pour trier les données à l'intérieur et les ranger où l'on souhaite
    tableur = load_workbook(filename = "listedescartes.xlsx")
    onglet = tableur.active
    reader = onglet.iter_rows(min_row = 1, max_row= 100, min_col=1, max_col=7)

    #Et on va pouvoir trier les lignes
    for row in reader:
        if row[1].value == "Commun" :
            newcarte = Cartestats(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value)
            commun.append(newcarte)
            listepaquetdecarte.append(newcarte.affiche)
        elif row[1].value == "Peu Commun":
            newcarte = Cartestats(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value)
            peucommun.append(newcarte)
            listepaquetdecarte.append(newcarte.affiche)
        elif row[1].value == "Rare":
            newcarte = Cartestats(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value)
            rare.append(newcarte)
            listepaquetdecarte.append(newcarte.affiche)
        elif row[1].value == "Unique":
            newcarte = Cartestats(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value)
            unique.append(newcarte)
            listepaquetdecarte.append(newcarte.affiche)
        else:
            continue
    
    #On ajoute les listes de cartes dans leur dico respectifs et on retourne 
    paquetdecarte["Commun"], paquetdecarte["Peu Commun"], paquetdecarte["Rare"], paquetdecarte["Unique"] = commun, peucommun, rare, unique
    return listepaquetdecarte

# if __name__ == __main__:
partie = Jeudecartestats()
partie.bataillesimple()
